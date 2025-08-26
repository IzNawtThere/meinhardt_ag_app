# parsers/excel_parser.py
"""
Professional Excel parser for Meinhardt Master Files
Based on the working parser with proper column mapping
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from typing import Dict, List, Tuple, Optional, Any
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

@dataclass
class DataPoint:
    code: str
    name: str
    pillar: str
    data_type: str
    display_name: str = None

@dataclass
class AssessmentCriteria:
    code: str
    name: str
    formula: str
    formula_type: str
    weight: float
    performance_signal_name: str
    data_points: List[str]
    thresholds: Dict[str, Any]

@dataclass
class PerformanceSignal:
    code: str
    name: str
    weight: float
    key_topic_name: str
    assessment_criteria: List[str]

@dataclass
class KeyTopic:
    code: str
    name: str
    pillar: str
    performance_signals: List[str]

class MasterFileParser:
    """Professional parser for Meinhardt Master Files"""
    
    # Expected pillar sheet names - match exactly or contains
    EXPECTED_PILLARS = {
        "P&M": "Planning & Monitoring",
        "D&T": "Design & Technical", 
        "D&C": "Development & Construction",
        "CE&O": "Cost Estimation & Optimization",
        "CEO": "Cost Estimation & Optimization",  # Alternative
        "I&T": "Innovation & Technology",
        "S&O": "Strategy & Operations"
    }
    
    def __init__(self, file_path: str):
        self.file_path = Path(file_path)
        if not self.file_path.exists():
            raise FileNotFoundError(f"Excel file not found: {file_path}")
        
        self.hierarchy = {
            'key_topics': {},
            'performance_signals': {},
            'assessment_criteria': {},
            'data_points': {}
        }
        self.relationships = {
            'kt_to_ps': {},
            'ps_to_ac': {},
            'ac_to_dp': {}
        }
        self.errors = []
        self.warnings = []
        
    def parse(self) -> Dict[str, Any]:
        """Main parsing function"""
        print(f"Starting parse of {self.file_path.name}")
        
        try:
            # Load workbook with openpyxl for merged cell support
            wb = load_workbook(self.file_path, data_only=True)
            print(f"Found sheets: {wb.sheetnames}")
            
            # Process each sheet
            for sheet_name in wb.sheetnames:
                pillar = self._match_pillar_name(sheet_name)
                if pillar:
                    print(f"Processing pillar: {pillar} (sheet: {sheet_name})")
                    self._parse_sheet_specific_format(wb[sheet_name], pillar, sheet_name)
                else:
                    # Skip known non-data sheets
                    skip_sheets = ['Ref', 'Reference', 'Summary', 'Instructions']
                    if not any(skip in sheet_name for skip in skip_sheets):
                        if 'QB' not in sheet_name:  # Skip QB sheets
                            print(f"Skipping non-pillar sheet: {sheet_name}")
            
            wb.close()
            
            print(f"\nParse complete:")
            print(f"  - {len(self.hierarchy['data_points'])} Data Points")
            print(f"  - {len(self.hierarchy['assessment_criteria'])} Assessment Criteria")
            print(f"  - {len(self.hierarchy['performance_signals'])} Performance Signals")
            print(f"  - {len(self.hierarchy['key_topics'])} Key Topics")
            
            return {
                'hierarchy': self.hierarchy,
                'relationships': self.relationships
            }
            
        except Exception as e:
            print(f"Error during parsing: {str(e)}")
            import traceback
            traceback.print_exc()
            raise
    
    def _match_pillar_name(self, sheet_name: str) -> Optional[str]:
        """Match sheet name to pillar"""
        sheet_upper = sheet_name.upper().strip()
        
        # Direct matching
        for key, pillar in self.EXPECTED_PILLARS.items():
            if key in sheet_upper:
                return pillar
        
        # Check if sheet contains pillar keywords
        if 'PLANNING' in sheet_upper or 'MONITORING' in sheet_upper:
            return "Planning & Monitoring"
        elif 'DESIGN' in sheet_upper or 'TECHNICAL' in sheet_upper:
            return "Design & Technical"
        elif 'DEVELOPMENT' in sheet_upper or 'CONSTRUCTION' in sheet_upper:
            return "Development & Construction"
        elif 'COST' in sheet_upper or 'ESTIMATION' in sheet_upper:
            return "Cost Estimation & Optimization"
        elif 'INNOVATION' in sheet_upper or 'TECHNOLOGY' in sheet_upper:
            return "Innovation & Technology"
        elif 'STRATEGY' in sheet_upper or 'OPERATIONS' in sheet_upper:
            return "Strategy & Operations"
        
        return None
    
    def _parse_sheet_specific_format(self, sheet, pillar: str, sheet_name: str):
        """
        Parse sheet with specific format:
        - Row 4-5: Headers
        - Row 6+: Data
        
        Column mapping:
        - Col B (2): Key Topic ID
        - Col C (3): Key Topic
        - Col D (4): Performance Signals ID
        - Col E (5): Performance Signals
        - Col F (6): Performance Signals Weightage
        - Col K (11): Data Point ID
        - Col L (12): Data Point
        - Col M (13): Assessment Criteria ID
        - Col N (14): Assessment Criteria
        - Col O (15): Assessment Criteria Weightage
        - Col P (16): Assessment Guidelines/Formula
        - Col Q (17): Good
        - Col R (18): Satisfactory
        - Col S (19): Needs Improvement
        """
        try:
            # Get merged cell ranges
            merged_ranges = list(sheet.merged_cells.ranges)
            merged_cell_map = {}
            
            for merged_range in merged_ranges:
                # Get value from top-left cell
                top_left_value = sheet.cell(
                    row=merged_range.min_row, 
                    column=merged_range.min_col
                ).value
                
                # Map all cells in range to this value
                for row in range(merged_range.min_row, merged_range.max_row + 1):
                    for col in range(merged_range.min_col, merged_range.max_col + 1):
                        merged_cell_map[(row, col)] = top_left_value
            
            # Track current context
            current_kt = None
            current_ps = None
            current_ac = None
            
            # Track last seen values to handle merged cells
            last_kt_name = None
            last_ps_name = None
            last_ac_name = None
            
            # Statistics for this sheet
            sheet_stats = {
                'key_topics': 0,
                'performance_signals': 0,
                'assessment_criteria': 0,
                'data_points': 0
            }
            
            # Process rows starting from row 6
            max_row = min(sheet.max_row + 1, 500)  # Limit for safety
            
            for row_num in range(6, max_row):
                # Helper function to get cell value
                def get_value(row, col):
                    if (row, col) in merged_cell_map:
                        return merged_cell_map[(row, col)]
                    val = sheet.cell(row=row, column=col).value
                    if val and isinstance(val, str) and val.startswith('='):
                        return None  # Skip formulas
                    return val
                
                # Extract values
                kt_id = get_value(row_num, 2)
                kt_name = get_value(row_num, 3)
                ps_id = get_value(row_num, 4)
                ps_name = get_value(row_num, 5)
                ps_weight = get_value(row_num, 6)
                dp_id = get_value(row_num, 11)
                dp_name = get_value(row_num, 12)
                ac_id = get_value(row_num, 13)
                ac_name = get_value(row_num, 14)
                ac_weight = get_value(row_num, 15)
                formula = get_value(row_num, 16)
                good = get_value(row_num, 17)
                satisfactory = get_value(row_num, 18)
                needs_improvement = get_value(row_num, 19)
                
                # Skip completely empty rows
                if all(v is None for v in [kt_name, ps_name, ac_name, dp_name]):
                    continue
                
                # Process Key Topic
                if kt_name and str(kt_name).strip() and kt_name != last_kt_name:
                    kt = KeyTopic(
                        code=str(kt_id).strip() if kt_id else f"KT-{sheet_stats['key_topics']+1}",
                        name=str(kt_name).strip(),
                        pillar=pillar,
                        performance_signals=[]
                    )
                    current_kt = kt.name
                    self.hierarchy['key_topics'][kt.name] = kt
                    last_kt_name = kt_name
                    sheet_stats['key_topics'] += 1
                
                # Process Performance Signal
                if ps_name and str(ps_name).strip() and ps_name != last_ps_name:
                    ps = PerformanceSignal(
                        code=str(ps_id).strip() if ps_id else f"PS-{sheet_stats['performance_signals']+1}",
                        name=str(ps_name).strip(),
                        weight=self._parse_weight(ps_weight),
                        key_topic_name=current_kt,
                        assessment_criteria=[]
                    )
                    current_ps = ps.name
                    self.hierarchy['performance_signals'][ps.name] = ps
                    last_ps_name = ps_name
                    sheet_stats['performance_signals'] += 1
                    
                    # Add relationship
                    if current_kt:
                        if current_kt not in self.relationships['kt_to_ps']:
                            self.relationships['kt_to_ps'][current_kt] = []
                        self.relationships['kt_to_ps'][current_kt].append(ps.name)
                        
                        # Add to KT's list
                        if current_kt in self.hierarchy['key_topics']:
                            self.hierarchy['key_topics'][current_kt].performance_signals.append(ps.name)
                
                # Process Assessment Criteria
                if ac_name and str(ac_name).strip() and ac_name != last_ac_name:
                    ac = AssessmentCriteria(
                        code=str(ac_id).strip() if ac_id else f"AC-{sheet_stats['assessment_criteria']+1}",
                        name=str(ac_name).strip(),
                        formula=str(formula).strip() if formula else "",
                        formula_type=self._determine_formula_type(str(formula) if formula else ""),
                        weight=self._parse_weight(ac_weight),
                        performance_signal_name=current_ps,
                        data_points=[],
                        thresholds={
                            'good': str(good).strip() if good else None,
                            'satisfactory': str(satisfactory).strip() if satisfactory else None,
                            'needs_improvement': str(needs_improvement).strip() if needs_improvement else None
                        }
                    )
                    current_ac = ac.name
                    self.hierarchy['assessment_criteria'][ac.name] = ac
                    last_ac_name = ac_name
                    sheet_stats['assessment_criteria'] += 1
                    
                    # Add relationship
                    if current_ps:
                        if current_ps not in self.relationships['ps_to_ac']:
                            self.relationships['ps_to_ac'][current_ps] = []
                        self.relationships['ps_to_ac'][current_ps].append(ac.name)
                        
                        # Add to PS's list
                        if current_ps in self.hierarchy['performance_signals']:
                            self.hierarchy['performance_signals'][current_ps].assessment_criteria.append(ac.name)
                
                # Process Data Point
                if dp_name and str(dp_name).strip() and current_ac:
                    # Don't add formulas as data points
                    if not str(dp_name).startswith('='):
                        dp = DataPoint(
                            code=str(dp_id).strip() if dp_id else f"DP-{sheet_stats['data_points']+1}",
                            name=str(dp_name).strip(),
                            pillar=pillar,
                            data_type=self._detect_data_type(str(dp_name))
                        )
                        
                        # Only add if not already exists (avoid duplicates)
                        if dp.name not in self.hierarchy['data_points']:
                            self.hierarchy['data_points'][dp.name] = dp
                            sheet_stats['data_points'] += 1
                        
                        # Add relationship
                        if current_ac:
                            if current_ac not in self.relationships['ac_to_dp']:
                                self.relationships['ac_to_dp'][current_ac] = []
                            if dp.name not in self.relationships['ac_to_dp'][current_ac]:
                                self.relationships['ac_to_dp'][current_ac].append(dp.name)
                            
                            # Add to AC's list
                            if current_ac in self.hierarchy['assessment_criteria']:
                                if dp.name not in self.hierarchy['assessment_criteria'][current_ac].data_points:
                                    self.hierarchy['assessment_criteria'][current_ac].data_points.append(dp.name)
            
            print(f"  Parsed {sheet_name}: {sheet_stats['key_topics']} KT, "
                  f"{sheet_stats['performance_signals']} PS, "
                  f"{sheet_stats['assessment_criteria']} AC, "
                  f"{sheet_stats['data_points']} DP")
            
        except Exception as e:
            print(f"Error parsing {sheet_name}: {str(e)}")
            import traceback
            traceback.print_exc()
    
    def _parse_weight(self, weight_value: Any) -> float:
        """Parse weight from various formats"""
        if weight_value is None:
            return 0.0
        
        # Skip formulas
        if isinstance(weight_value, str) and weight_value.startswith('='):
            return 0.0
        
        try:
            weight_str = str(weight_value).strip()
            
            # Remove % sign
            if '%' in weight_str:
                weight_str = weight_str.replace('%', '').strip()
                return float(weight_str)
            
            # Direct conversion
            weight = float(weight_str)
            
            # If between 0 and 1, convert to percentage
            if 0 < weight <= 1:
                return weight * 100
            
            return weight
            
        except:
            return 0.0
    
    def _detect_data_type(self, dp_name: str) -> str:
        """Detect data type from name"""
        if not dp_name:
            return 'text'
        
        name_lower = dp_name.lower()
        
        # Check explicit indicators
        if '(no.)' in name_lower or '(number)' in name_lower:
            return 'number'
        elif '(%)' in name_lower or 'percentage' in name_lower:
            return 'percentage'
        elif '(dd/mm/yy)' in name_lower or '(date)' in name_lower:
            return 'date'
        elif '(yes/no)' in name_lower:
            return 'boolean'
        
        # Check patterns
        if any(word in name_lower for word in ['value', 'cost', 'budget', 'amount', 'index', 'score']):
            return 'number'
        elif any(word in name_lower for word in ['rate', 'percent', '%']):
            return 'percentage'
        elif any(word in name_lower for word in ['date', 'time', 'deadline', 'schedule']):
            return 'date'
        elif any(word in name_lower for word in ['yes', 'no', 'true', 'false', 'is ', 'has ', 'does ']):
            return 'boolean'
        
        return 'text'
    
    def _determine_formula_type(self, formula: str) -> str:
        """Determine if formula is quantitative or qualitative"""
        if not formula:
            return 'qualitative'
        
        math_operators = ['+', '-', '*', '/', '%', '(', ')']
        if any(op in formula for op in math_operators):
            return 'quantitative'
        return 'qualitative'