def clean_headers(cols):
    counts = {}
    cleaned = []
    for col in cols:
        try:
            base = str(col).replace("\n", " ").strip().lower().replace(" ", "_")
        except:
            base = "unnamed"
        base = "unnamed" if base in [None, "", "nan"] else base
        if base in counts:
            counts[base] += 1
            base = f"{base}_{counts[base]}"
        else:
            counts[base] = 0
        cleaned.append(base)
    return cleaned

import re

def extract_field_type(description):
    description = str(description).lower()

    if "no." in description or "number" in description:
        return "No."
    elif "%" in description or "percent" in description:
        return "%"
    elif "dd/mm/yy" in description or "date" in description:
        return "dd/mm/yy"
    else:
        return "Text"

import pandas as pd

def extract_cleaned_df(df_raw):
    # Find header row dynamically based on 'data point'
    header_row_idx = None
    for i in range(len(df_raw)):
        row = df_raw.iloc[i].astype(str).str.lower()
        if row.str.contains("data point").any():
            header_row_idx = i
            break

    if header_row_idx is None:
        raise ValueError("No header row containing 'data point' found.")

    df_raw.columns = df_raw.iloc[header_row_idx]
    df_data = df_raw.iloc[header_row_idx + 1:].reset_index(drop=True)
    df_data.columns = clean_headers(df_data.columns)
    df_data.dropna(axis=1, how="all", inplace=True)
    return df_data

def ensure_columns(df):
    required_cols = ["section_id", "description", "default_value", "data_point_id", "devco_id"]
    for col in required_cols:
        if col not in df.columns:
            df[col] = None

import openpyxl
import pandas as pd
from openpyxl.utils import get_column_letter

def extract_dataframe_from_sheet(ws):
    # Step 1: Flatten merged cells
    merged_value_map = {}
    for merged_range in ws.merged_cells.ranges:
        top_left_value = ws.cell(merged_range.min_row, merged_range.min_col).value
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                coord = f"{get_column_letter(col)}{row}"
                merged_value_map[coord] = top_left_value

    # Step 2: Extract values row-wise
    data_matrix = []
    for row in ws.iter_rows():
        row_data = []
        for cell in row:
            coord = cell.coordinate
            value = merged_value_map.get(coord, cell.value)
            row_data.append(value)
        data_matrix.append(row_data)

    # Step 3: Convert to DataFrame
    df_raw = pd.DataFrame(data_matrix)
    return df_raw